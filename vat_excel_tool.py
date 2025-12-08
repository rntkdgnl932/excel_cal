# vat_excel_tool.py
# 하비 브라운 전용: 부가세/할인 계산 + 엑셀 템플릿 채우기 (최종 수정: 기존 데이터 청소 기능 강화)



from dataclasses import dataclass
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.cell import MergedCell as _MC
from openpyxl.styles import Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

# --------------------- 공통 헬퍼 --------------------- #

def _safe_cells(ws: Worksheet):
    """MergedCell 은 건너뛰고 실제 Cell 만 yield."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, _MC):
                continue
            yield cell


def _set_value_right_of_label(ws: Worksheet, label: str, value):
    """
    시트에서 'label' 텍스트를 가진 셀을 찾고,
    그 오른쪽으로 가면서 처음 만나는 '머지 안 됐고 비어 있는 셀'에 value 를 넣는다.
    """
    for cell in _safe_cells(ws):
        if str(cell.value).strip() == label:
            row = cell.row
            for col in range(cell.column + 1, ws.max_column + 1):
                target = ws.cell(row=row, column=col)
                if isinstance(target, _MC):
                    continue
                if target.value is None or str(target.value).strip() == "":
                    target.value = value
                    return


def _replace_exact_text(ws: Worksheet, old: str, new: str):
    """
    셀 값이 old 와 정확히 같은 경우만 new 로 대체.
    """
    for cell in _safe_cells(ws):
        if str(cell.value).strip() == old:
            cell.value = new


def _normalize(text: Optional[str]) -> str:
    if text is None:
        return ""
    return str(text).replace(" ", "").strip()


def _format_iso_to_kr(date_str: str) -> str:
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return f"{dt.year}년 {dt.month}월 {dt.day}일"
    except Exception:
        return date_str


# ------------- 숫자 → 한글 금액(견적금액용) ------------- #

_KR_NUM = ["", "일", "이", "삼", "사", "오", "육", "칠", "팔", "구"]
_KR_UNIT_SMALL = ["", "십", "백", "천"]
_KR_UNIT_BIG = ["", "만", "억", "조", "경"]


def _int_to_korean_amount(n) -> str:
    n = int(float(n))

    if n == 0:
        return "영"
    if n < 0:
        return "마이너스 " + _int_to_korean_amount(-n)

    parts: List[str] = []
    unit_pos = 0

    while n > 0:
        four = n % 10000
        n //= 10000

        if four == 0:
            unit_pos += 1
            continue

        small_parts: List[str] = []
        for i in range(4):
            digit = four % 10
            four //= 10
            if digit != 0:
                small = ""
                if not (digit == 1 and i > 0):
                    small += _KR_NUM[digit]
                small += _KR_UNIT_SMALL[i]
                small_parts.append(small)
        small_str = "".join(reversed(small_parts))

        if _KR_UNIT_BIG[unit_pos]:
            small_str += _KR_UNIT_BIG[unit_pos]

        parts.append(small_str)
        unit_pos += 1

    return " ".join(reversed(parts)).strip()


# ---------------------------------------------------------------------------
# 데이터 구조
# ---------------------------------------------------------------------------

@dataclass
class TradeInfo:
    customer_name: str
    supply_date: str
    biz_no: str
    contact: str
    vat_rate: float


@dataclass
class LineItemInput:
    name: str
    spec: str
    qty: int
    unit_gross: int
    discount_rate: float


@dataclass
class LineItemComputed(LineItemInput):
    unit_supply_original: int          # 할인 전 1개당 공급가
    unit_discounted_gross: int         # 할인 후 1개당 (부가세 포함) 금액
    unit_supply_discounted: int        # 할인 후 1개당 공급가(부가세 제외)
    unit_vat: int                      # ✅ 할인 후 1개당 부가세
    supply_total: int                  # 전체 공급가(합계)
    vat_total: int                     # 전체 부가세(합계)
    gross_total: int                   # 전체 합계(공급가+부가세)


HeaderMap = Dict[str, int]


# ---------------------------------------------------------------------------
# 계산 로직
# ---------------------------------------------------------------------------

def compute_items_with_vat(
        items: List[LineItemInput], vat_rate: float
) -> List[LineItemComputed]:
    result: List[LineItemComputed] = []
    rate_vat = vat_rate / 100.0

    for it in items:
        unit_supply_original = round(it.unit_gross / (1 + rate_vat))
        unit_discounted_gross = round(it.unit_gross * (1 - it.discount_rate / 100.0))
        unit_supply_discounted = round(unit_discounted_gross / (1 + rate_vat))
        unit_vat = unit_discounted_gross - unit_supply_discounted   # ✅ 1개당 부가세

        gross_total = unit_discounted_gross * it.qty
        supply_total = round(gross_total / (1 + rate_vat))
        vat_total = gross_total - supply_total

        result.append(
            LineItemComputed(
                name=it.name,
                spec=it.spec,
                qty=it.qty,
                unit_gross=it.unit_gross,
                discount_rate=it.discount_rate,
                unit_supply_original=unit_supply_original,
                unit_discounted_gross=unit_discounted_gross,
                unit_supply_discounted=unit_supply_discounted,
                unit_vat=unit_vat,                    # ✅ 추가
                supply_total=supply_total,
                vat_total=vat_total,
                gross_total=gross_total,
            )
        )
    return result



# ---------------------------------------------------------------------------
# 엑셀 유틸 (본문/합계)
# ---------------------------------------------------------------------------

def _find_detail_header(ws: Worksheet) -> Tuple[int, HeaderMap]:
    header_row_idx = -1
    col_map: HeaderMap = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        texts = [_normalize(c.value) for c in row]
        if not texts:
            continue

        has_item_col = any(("품목" in t) or ("품명" in t) for t in texts)

        if has_item_col:
            header_row_idx = row[0].row
            for idx, text in enumerate(texts, start=1):
                # 1. 순번
                if text in ("NO", "No", "순번", "번호"):
                    col_map["seq"] = idx
                # 2. 품명/규격/단위/수량
                elif ("품목" in text) or ("품명" in text):
                    col_map["name"] = idx
                elif "규격" in text:
                    col_map["spec"] = idx
                elif "단위" in text:
                    col_map["unit"] = idx
                elif "수량" in text:
                    col_map["qty"] = idx
                # 3. 단가
                elif text.startswith("단가"):
                    col_map["unit_price"] = idx
                # 4. 공급가/부가세
                elif text in ("금액", "공급가액", "공급가"):
                    col_map["supply"] = idx
                elif text in ("세액", "부가세"):
                    col_map["vat"] = idx

                # [수정 핵심] 합계가 이미 찾아졌으면 비고는 무시함
                elif text in ("합계", "총액", "합계금액"):
                    col_map["gross"] = idx
                elif "비고" in text:
                    if "gross" not in col_map:
                        col_map["gross"] = idx
            break

    if header_row_idx == -1:
        raise RuntimeError("시트에서 '품명/품목' 헤더 행을 찾지 못했습니다.")

    return header_row_idx, col_map


def _is_merged(cell) -> bool:
    return isinstance(cell, MergedCell)


def _clear_body(ws, start_row: int):
    """본문 데이터 영역을 청소합니다."""
    row = start_row
    while True:
        if row > 500: break

        is_footer = False
        for col in range(1, 16):
            cell = ws.cell(row=row, column=col)
            val = str(cell.value).replace(" ", "") if cell.value else ""
            if any(x in val for x in ["총합계금액", "소계", "부가세", "합계"]):
                is_footer = True
                break
        if is_footer: break

        for col in range(1, 30):  # 컬럼 범위 넉넉하게
            cell = ws.cell(row=row, column=col)
            if not isinstance(cell, _MC):
                cell.value = None
        row += 1


def _write_items_to_sheet(ws, items: List[LineItemComputed]) -> Tuple[int, HeaderMap]:
    header_row, col_map = _find_detail_header(ws)
    body_start = header_row + 1

    _clear_body(ws, body_start)

    # [수정] 모든 테두리를 얇은 실선으로 만드는 스타일 정의
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for idx, item in enumerate(items, start=1):
        row = body_start + (idx - 1)

        def set_if(col_key: str, value):
            col = col_map.get(col_key)
            if col is None:
                return
            cell = ws.cell(row=row, column=col)
            if _is_merged(cell):
                return

            cell.value = value

            # 숫자 포맷 적용
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0'

            # [핵심 수정] 여기서 테두리를 강제로 얇게 설정합니다.
            # 이중선(double)이 있어도 덮어씌워집니다.
            cell.border = thin_border
            # 글자 정렬도 깔끔하게 (가운데 정렬) - 필요 없으면 주석 처리
            # cell.alignment = Alignment(horizontal='center', vertical='center')

        set_if("seq", idx)
        set_if("name", item.name)
        set_if("spec", item.spec)
        set_if("unit", "EA")
        set_if("qty", item.qty)
        # 단가/공급가/부가세를 1개당 기준으로 표시
        set_if("unit_price", item.unit_supply_discounted)  # (그대로 사용)
        set_if("supply", item.unit_supply_discounted)  # ✅ 1개당 공급가
        set_if("vat", item.unit_vat)  # ✅ 1개당 부가세

        # 합계금액은 기존처럼 전체 금액
        set_if("gross", item.gross_total)

    return header_row, col_map


def _calc_totals(items: List[LineItemComputed]) -> Tuple[int, int, int]:
    total_supply = sum(it.supply_total for it in items)
    total_vat = sum(it.vat_total for it in items)
    total_gross = sum(it.gross_total for it in items)
    return total_supply, total_vat, total_gross


def _fill_footer_totals_common(ws: Worksheet, items: List[LineItemComputed]) -> None:
    total_supply, total_vat, total_gross = _calc_totals(items)
    _set_value_right_of_label(ws, "소계", total_supply)
    _set_value_right_of_label(ws, "부가세", total_vat)
    _set_value_right_of_label(ws, "총합계금액", total_gross)


# ---------------------------------------------------------------------------
# 헤더(거래처/날짜/견적금액) 채우기
# ---------------------------------------------------------------------------

def _fill_common_replace(ws, info: TradeInfo) -> None:
    customer_name = getattr(info, "customer_name", "") or getattr(info, "customer", "")
    if not customer_name:
        return
    _replace_exact_text(ws, "거래처명", customer_name)


def _fill_dates(ws, info: TradeInfo):
    date_labels = ["견적일자", "공급일자", "공급일", "납품일"]

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            t = _normalize(v)
            target_label = None
            for lbl in date_labels:
                if lbl in t:
                    target_label = lbl
                    break
            if not target_label:
                continue

            if "납품일" in target_label:
                date_value = _format_iso_to_kr(info.supply_date)
            else:
                date_value = info.supply_date

            r = cell.row
            for c in range(cell.column + 1, ws.max_column + 1):
                target = ws.cell(row=r, column=c)
                if _is_merged(target):
                    continue
                if isinstance(target.value, str) and target.value.strip():
                    continue
                target.value = date_value
                break


# [추가] 거래명세표처럼 라벨 없이 "0000년 00월 00일" 형태를 찾아 바꾸는 함수
def _fill_korean_style_date(ws: Worksheet, date_str: str):
    """
    YYYY-MM-DD 문자열을 받아 상단 영역(1~10행)에서
    '년', '월', '일' 글자가 포함된 셀 근처의 숫자를 찾아 날짜를 업데이트합니다.
    """
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        year, month, day = dt.year, dt.month, dt.day
    except:
        return  # 날짜 형식이 아니면 패스

    # 상단 10줄만 탐색
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            if isinstance(cell, _MC) or cell.value is None:
                continue

            val = str(cell.value).strip()

            # 1. 셀 자체가 "2024년" 처럼 되어 있는 경우
            if "년" in val and "월" not in val:  # 년도만 있는 셀
                # 숫자만 추출해서 교체 시도
                new_val = val.replace(str(datetime.now().year), str(year))  # 현재 년도 -> 입력 년도
                # 만약 못 찾았으면 그냥 통째로 "2025년"으로 교체 시도
                # (정규식 등으로 더 정교하게 할 수 있으나 단순하게 처리)
                if str(year) not in new_val:
                    # 기존 값에 숫자가 있으면 그 숫자를 새 년도로 교체
                    import re
                    new_val = re.sub(r'\d+', str(year), val)
                cell.value = new_val

            elif "월" in val and "일" not in val and "년" not in val:  # 월만 있는 셀
                import re
                cell.value = re.sub(r'\d+', str(month), val)

            elif "일" in val and "월" not in val:  # 일만 있는 셀
                import re
                cell.value = re.sub(r'\d+', str(day), val)

            # 2. "2025년 8월 20일" 처럼 한 셀에 다 들어있는 경우
            elif "년" in val and "월" in val and "일" in val:
                cell.value = f"{year}년 {month}월 {day}일"

            # 3. (옵션) 셀에는 "년"이라고만 적혀있고, 바로 왼쪽 셀이 숫자인 경우 (템플릿 구조에 따라 다름)
            # 이 경우는 복잡하므로 위 1, 2번 케이스로 대부분 해결될 것입니다.


def _get_writable_cell(ws: Worksheet, coord: str):
    """
    병합된 셀 범위 안의 좌표가 들어오면,
    항상 '왼쪽 위 대표 셀'을 돌려준다.
    (MergedCell 에 직접 쓰다가 나는 에러 방지용)
    """
    cell = ws[coord]
    if not isinstance(cell, MergedCell):
        return cell

    # 병합 범위들 중에서 이 좌표가 들어있는 범위를 찾는다.
    for mr in ws.merged_cells.ranges:
        if coord in mr:
            return ws.cell(row=mr.min_row, column=mr.min_col)

    # 혹시 못 찾으면 그냥 원래 셀 그대로 반환
    return cell

def _fill_quote_total(ws, items: List[LineItemComputed]) -> None:
    total = int(sum(it.gross_total for it in items))

    # "견적금액"이 있는 행 찾기 그대로 유지
    label_row = None
    for row in ws.iter_rows(min_row=1, max_row=30):
        for cell in row:
            if isinstance(cell.value, str) and "견적금액" in _normalize(cell.value):
                label_row = cell.row
                break
        if label_row is not None:
            break

    if label_row is None:
        return

    row = label_row

    # 병합 대응 헬퍼로 숫자 / 한글 셀 가져오기
    cell_number = _get_writable_cell(ws, f"H{row}")   # 합계 숫자 들어가는 칸
    cell_korean = _get_writable_cell(ws, f"M{row}")   # 한글 "( ... )" 들어가는 칸

    # --- 여기부터가 핵심 수정 ---

    # 1) 숫자 셀: 값 + 서식에 ₩ 포함
    cell_number.value = total
    cell_number.number_format = '"₩" #,##0'   # => ₩ 7,040,000

    # 2) 한글 셀: 너가 바꾼 형식 유지
    korean_money = _int_to_korean_amount(total)
    cell_korean.value = f" {korean_money} "
    cell_korean.alignment = Alignment(horizontal="center", vertical="center")





def _fill_statement_totals(ws: Worksheet, items: List[LineItemComputed]) -> None:
    total_supply, total_vat, total_gross = _calc_totals(items)

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, _MC):
                continue
            v = _normalize(cell.value)
            target_val = None

            if "소계" in v:
                target_val = total_supply
            elif "부가세" in v:
                target_val = total_vat
            elif "총합계금액" in v or "합계금액" in v:
                target_val = total_gross

            if target_val is not None:
                r = cell.row
                for c in range(cell.column + 1, ws.max_column + 1):
                    target_cell = ws.cell(row=r, column=c)
                    if isinstance(target_cell, _MC):
                        continue
                    target_cell.value = target_val
                    break


def _fill_delivery_totals(
        ws: Worksheet, items: List[LineItemComputed]
) -> None:
    total_supply, total_vat, total_gross = _calc_totals(items)
    header_row, col_map = _find_detail_header(ws)

    sum_row = None
    for r in range(header_row + 1, ws.max_row + 1):
        row_values = ""
        for c in range(1, 10):
            v = ws.cell(row=r, column=c).value
            if v: row_values += str(v)

        norm_val = _normalize(row_values)
        if "합계" in norm_val or "총계" in norm_val or "Total" in norm_val:
            sum_row = r
            break

    if sum_row is None:
        sum_row = header_row + len(items) + 1

    def set_cell(col_key: str, value: int):
        col = col_map.get(col_key)
        if not col:
            return
        target = ws.cell(row=sum_row, column=col)
        if isinstance(target, MergedCell):
            return
        target.value = value
        target.number_format = '#,##0'

    set_cell("supply", total_supply)
    set_cell("vat", total_vat)
    set_cell("gross", total_gross)


# ---------------------------------------------------------------------------
# 템플릿별 진입 함수
# ---------------------------------------------------------------------------



def _fill_quote_header_dates(ws: Worksheet) -> None:
    """견적서 상단의 '견적번호'와 '견적일자'를 오늘 기준으로 채웁니다.
    - 견적번호: HHMMSS
    - 견적일자: YYYY-MM-DD
    템플릿에서 해당 라벨 오른쪽의 첫 번째 비-병합 셀에 값을 씁니다.
    """
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")   # {연}-{월}-{일}
    num_str = now.strftime("%H%M%S")      # {시}{분}{초}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            v = cell.value
            if not isinstance(v, str):
                continue
            t = _normalize(v)
            r = cell.row

            # 견적일자 → 오늘 날짜
            if "견적일자" in t:
                for c in range(cell.column + 1, ws.max_column + 1):
                    target = ws.cell(row=r, column=c)
                    if _is_merged(target):
                        continue
                    # 기존 값이 있어도 오늘 날짜로 덮어씀
                    target.value = date_str
                    break

            # 견적번호 → HHMMSS
            elif "견적번호" in t:
                for c in range(cell.column + 1, ws.max_column + 1):
                    target = ws.cell(row=r, column=c)
                    if _is_merged(target):
                        continue
                    target.value = num_str
                    break

def fill_quote_template(
        template_path: Path,
        output_path: Path,
        info: TradeInfo,
        items: List[LineItemComputed],
) -> None:
    wb = load_workbook(template_path)
    ws = wb.active

    _fill_common_replace(ws, info)
    _fill_dates(ws, info)              # 공급일자/납품일 등은 기존 로직 유지
    _fill_quote_header_dates(ws)       # ✅ 견적번호/견적일자만 오늘 기준으로 덮어쓰기

    _set_value_right_of_label(ws, "납품장소", info.customer_name)
    _set_value_right_of_label(ws, "납기일자", "시안 확정 후 영업일 기준 10일 내외")

    _write_items_to_sheet(ws, items)
    _fill_quote_total(ws, items)
    _fill_footer_totals_common(ws, items)

    wb.save(output_path)



def fill_delivery_template(
        template_path: Path,
        output_path: Path,
        info: TradeInfo,
        items: List[LineItemComputed],
) -> None:
    wb = load_workbook(template_path)
    ws = wb.active

    _fill_common_replace(ws, info)
    _fill_dates(ws, info)

    _set_value_right_of_label(ws, "사업장소재지", info.customer_name)
    _set_value_right_of_label(ws, "공급받는자", info.customer_name)

    _write_items_to_sheet(ws, items)
    _fill_delivery_totals(ws, items)

    wb.save(output_path)


def fill_statement_template(
        template_path: Path,
        output_path: Path,
        info: TradeInfo,
        items: List[LineItemComputed],
) -> None:
    wb = load_workbook(template_path)
    ws = wb.active

    _fill_common_replace(ws, info)

    # 기존 날짜 함수 (라벨 찾는 방식) - 혹시 모르니 유지
    _fill_dates(ws, info)

    # [추가] 라벨 없이 년/월/일 글자 찾아서 날짜 바꾸는 함수 실행
    _fill_korean_style_date(ws, info.supply_date)

    _write_items_to_sheet(ws, items)
    _fill_statement_totals(ws, items)

    wb.save(output_path)







